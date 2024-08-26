import tkinter as tk
from tkinter import messagebox, ttk
from tkinter import font
from tkcalendar import DateEntry
import sql  # sql.py 파일을 import
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import json

def save_settings(settings, filename="settings.json"):
    with open(filename, "w") as f:
        json.dump(settings, f)

def load_settings(filename="settings.json"):
    try:
        with open(filename, "r") as f:
            settings = json.load(f)
    except FileNotFoundError:
        # 기본 설정을 반환 (파일이 없을 경우)
        settings = {"main_font_size": 100, "main_window_size": "1400x800", "edit_font_size": 100, "edit_window_size": "600x800", "sessions_font_size": 100, "sessions_window_size": "1000x600"}
    return settings

# GUI 애플리케이션 생성
root = tk.Tk()
root.title("김경희 심리상담 센터")
settings = load_settings()
root.geometry(settings["main_window_size"])  # 메인 창 초기 크기 설정

# 기본 폰트 크기
BASE_FONT_SIZE = 10
custom_font = ("맑은 고딕", BASE_FONT_SIZE)

# Treeview 설정
headers = [
    "ID", "이름", "생년월일", "년", "월", "일", "나이", "메일", "성별", "남자", "여자", "전화번호", "주소", "상담시작일", "시작연도", "시작월", "시작일", "상담종료일", "종료연도", "종료월", "종료일", "호소문제", "회기 수", "특이사항"
]
viewonly_headers = [
    "ID", "이름", "생년월일", "나이", "메일", "성별", "전화번호", "주소", "상담시작일", "상담종료일", "회기 수"
]
export_headers = [
    "ID", "이름", "생년월일", "나이", "메일", "성별", "전화번호", "주소", "상담시작일", "상담종료일", "호소문제", "회기 수", "특이사항"
]
export_detail_headers = [
    "회차", "날짜", "상담내용"
]

treeview_users = ttk.Treeview(root, columns=viewonly_headers, show='headings')

# 기본 열 너비 설정
column_widths = {
    "ID": 10,
    "이름": 50,
    "생년월일": 100,
    "나이": 10,
    "메일": 150,
    "성별": 10,
    "전화번호": 100,
    "주소": 300,
    "상담시작일": 100,
    "상담종료일": 100,
    "회기 수": 10
}

for header in viewonly_headers:
    treeview_users.heading(header, text=header)
    treeview_users.column(header, width=column_widths.get(header, 100), anchor=tk.CENTER)

treeview_users.grid(row=2, column=0, columnspan=5, padx=10, pady=10, sticky='nsew')

gender_var = tk.StringVar()

main_bar = 1
edit_bar = 2
sessions_bar = 3
updated_row = 1

def on_mouse_wheel(event, canvas, scrollable_frame):
    # 마우스 휠의 이동 방향에 따라 스크롤을 조정합니다.
    if scrollable_frame.winfo_containing(event.x_root, event.y_root) == scrollable_frame:
        if event.delta > 0:  # 위로 스크롤
            canvas.yview_scroll(-1, "units")
        else:  # 아래로 스크롤
            canvas.yview_scroll(1, "units")

def add_scrollbar_to_window(root):
    # Canvas와 Scrollbar를 생성합니다.
    canvas = tk.Canvas(root)
    canvas.grid(row=0, column=0, sticky="nsew")
    
    v_scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    v_scrollbar.grid(row=0, column=1, sticky="ns")

    h_scrollbar = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)
    h_scrollbar.grid(row=1, column=0, sticky="ew")
    
    scrollable_frame = tk.Frame(canvas)

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # 전체 윈도우에서 마우스 휠 스크롤을 감지하도록 설정
    root.bind_all("<MouseWheel>", lambda e, c=canvas, sw=scrollable_frame: on_mouse_wheel(e, c, sw))

    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    return scrollable_frame

def update_window_size(percent, bar_type=None):
    global updated_row
    if bar_type == main_bar:
        style = ttk.Style()
        style.configure("Treeview", font=custom_font)
        text_width = 14 * percent
        text_height = 8 * percent
        windw_name = "main_window_size"
        
    elif bar_type == edit_bar:
        text_width = 6 * percent
        text_height = 8 * percent
        windw_name = "edit_window_size"

    elif bar_type == sessions_bar:
        if percent == 100:
            x_offset = 10
            y_offset = 18
        elif percent ==110:
            x_offset = 20
            y_offset = 19
        elif percent ==120:
            x_offset = 30
            y_offset = 20
        elif percent ==130:
            x_offset = -40
            y_offset = 21
        elif percent ==140:
            x_offset = 50
            y_offset = 22
        else:
            x_offset = -20
            y_offset = 23

        text_width = 10 * percent + x_offset
        if updated_row  >= y_offset:
            text_height = 6 * percent + ((updated_row - (y_offset-1)) * 30)
        else:
            text_height = 6 * percent
        
        windw_name = "sessions_window_size"
    
    window_size = f"{text_width}x{text_height}"
    settings[windw_name] = window_size
    save_settings(settings)

def update_font_size(percent, all_widgets, bar_type=None):
    size = int(BASE_FONT_SIZE * percent / 100)
    global custom_font, settings
    custom_font = ("맑은 고딕", size)
    
    # 적용할 위젯 리스트
    for widget in all_widgets:
        widget.config(font=custom_font)
    
    settings = load_settings()

    # Treeview의 스타일 업데이트
    if bar_type == main_bar:
        style = ttk.Style()
        style.configure("Treeview", font=custom_font)
        settings["main_font_size"] = percent
    elif bar_type == edit_bar:
        settings["edit_font_size"] = percent
    elif bar_type == sessions_bar:
        settings["sessions_font_size"] = percent
    
    save_settings(settings)

def set_font_size(percent, all_widgets, bar_type):
    update_font_size(percent, all_widgets, bar_type)
    update_window_size(percent, bar_type)

def create_menu(root, all_widgets, bar_type):
    global font_menu, percent_buttons, var
    menu_bar = tk.Menu(root)
    root.config(menu=menu_bar)
    
    settings_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="추가기능", menu=settings_menu)
    
    font_menu = tk.Menu(settings_menu, tearoff=0)
    settings_menu.add_cascade(label="글꼴", menu=font_menu)

    # 퍼센트 항목 추가
    percent_buttons = [100, 110, 120, 130, 140, 150]

    # 현재 선택된 폰트 크기를 저장할 변수
    if bar_type == main_bar:
        var = tk.IntVar(value=settings.get("main_font_size", 100))
    elif bar_type == edit_bar:
        var = tk.IntVar(value=settings.get("edit_font_size", 100))
    elif bar_type == sessions_bar:
        var = tk.IntVar(value=settings.get("sessions_font_size", 100))
    
    for percent in percent_buttons:
        font_menu.add_radiobutton(label=f"{percent}%", variable=var, value=percent, command=lambda p=percent: set_font_size(p, all_widgets, bar_type))

def update_age(year, month, day):
    today = datetime.today()
    age = today.year - year - ((today.month, today.day) < (month, day))
    return age

def validate_integer(input_value):
    """정수만 허용하는 검증 함수"""
    return input_value.isdigit() or input_value == ""

def validate_phone(input_value):
    """전화번호에 숫자와 공백만 허용하는 검증 함수"""
    return all(char.isdigit() or char.isspace() for char in input_value) or input_value == ""

def validate_hyphen(input_value):
    """"-"만 허용하는 검증 함수"""
    return  input_value!="-"

def update_sessions_detail(user_id, session_number, session_date, details):
    # 데이터가 이미 존재하는지 확인
    result = sql.check_session_detail(user_id, session_number)
    
    if result:
        # 데이터가 존재하면 업데이트
        detail_id = result[0]
        sql.update_session_detail(detail_id, session_date, details)
    else:
        sql.create_session_detail(user_id, session_number, session_date, details)

slash_type = 0
hyphen_type = 1

def convert_date_format(date_str, date_type):
    try:
        if date_type == hyphen_type:
            date_obj = datetime.strptime(date_str, '%m/%d/%y')
            converted_date = date_obj.strftime('%Y-%m-%d')
        else:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            converted_date = date_obj.strftime('%m/%d/%y')
        return converted_date
    except ValueError:
        return date_str

# 세션 저장 함수
def save_session_data(edit_window, session_window, session_data, user_id, user_data):
    # 데이터베이스에서 해당 세션 삭제
    last_row_num = len(session_data)
    exist_row_num = len(user_data)
    
    if last_row_num < exist_row_num:
        for i in range(last_row_num, exist_row_num):
            last_detail_id = user_data[i][0]
            sql.delete_session_detail(last_detail_id)
    
    for row_num, entries in session_data.items():
        session_number = row_num
        session_date = convert_date_format(entries[0], hyphen_type)
        details = entries[1]
        
        update_sessions_detail(user_id, session_number, session_date, details)
    
    messagebox.showinfo("저장 완료", "세션 정보가 저장되었습니다.")
    session_window.destroy()

    edit_window.deiconify()

def add_entry_row(session_window, labels, session_data, row_num, add_button, save_button, delete_button, preset_data=None, session_window_widgets=None):
    entries = []
    global updated_row

    # '회차' 라벨 추가
    label = tk.Label(session_window, text=f"{row_num}회기")
    label.grid(row=row_num, column=0, padx=10, pady=5)

    if session_window_widgets is not None:
        session_window_widgets.append(label)

    for i, label_text in enumerate(labels[1:], start=1):
        if label_text == "날짜":
            entry = DateEntry(session_window, width=12, background='darkblue', foreground='white', borderwidth=2)
            if preset_data:
                entry.set_date(preset_data[0])  # 날짜를 미리 채웁니다.
            entry.grid(row=row_num, column=i, padx=10, pady=5)
        else:
            entry = tk.Entry(session_window, width=80)
            if preset_data:
                entry.insert(0, preset_data[1])  # 상담 내용을 미리 채웁니다.
            entry.grid(row=row_num, column=i, padx=10, pady=5)
        entries.append(entry)
        if session_window_widgets is not None:
            session_window_widgets.append(entry)
    
    # 각 행의 입력 필드를 session_data 리스트에 추가
    session_data[row_num] = entries
    
    # '추가', '저장', '삭제' 버튼 위치 재조정
    add_button.grid(row=row_num + 1, column=len(labels) + 1, padx=5, pady=5)
    delete_button.grid(row=row_num + 1, column=len(labels) + 2, padx=5, pady=5)
    save_button.grid(row=row_num + 1, column=len(labels) + 3, padx=5, pady=10)

    if session_window_widgets is not None:
        session_window_widgets.extend([add_button, delete_button, save_button])
    
    updated_row = row_num + 2
    update_font_size(settings["sessions_font_size"], session_window_widgets, sessions_bar)
    update_window_size(settings["sessions_font_size"], sessions_bar)

# 입력 행 추가 함수
def delete_entry_row(session_window, labels, session_data, add_button, save_button, delete_button, session_window_widgets):
    global updated_row

    if len(session_data) > 1:
        # 마지막 입력 행 삭제
        last_row_num = max(session_data.keys())
        for entry in session_data[last_row_num]:
            entry.destroy()
        
        session_window.grid_slaves(row=last_row_num)[0].destroy()  # 회차 라벨 제거
        del session_data[last_row_num]

        start_Index = 9 + 6 * (last_row_num - 2)
        end_Indx = start_Index + 6
        del session_window_widgets[start_Index:end_Indx]
        
        # '추가', '저장', '삭제' 버튼 위치 재조정
        add_button.grid(row=len(session_data) + 1, column=len(labels) + 1, padx=5, pady=5)
        delete_button.grid(row=len(session_data) + 1, column=len(labels) + 2, padx=5, pady=5)
        save_button.grid(row=len(session_data) + 1, column=len(labels) + 3, padx=5, pady=10)
        
        updated_row = len(session_data) + 2
        update_window_size(settings["sessions_font_size"], sessions_bar)

# 회기 세부 정보 입력 창을 여는 함수
def open_sessions(window, user_id=None):
    settings = load_settings()

    session_window = tk.Toplevel(window)
    session_window.title("회기 세부 정보")
    session_window.geometry(settings["sessions_window_size"])  # 새 창 크기 설정
    
    # 스크롤 가능한 Frame 생성
    scrollable_frame = add_scrollbar_to_window(session_window)
    
    # 세션 데이터 저장할 딕셔너리
    session_data = {}
    session_window_widgets = []

    # 테이블 헤더 생성
    labels = ["회차", "날짜", "상담내용"]
    for i, label_text in enumerate(labels):
        label = tk.Label(scrollable_frame , text=label_text)
        label.grid(row=0, column=i, padx=10, pady=5)
        session_window_widgets.append(label)
    
    user_data =sql.get_session_details_by_user(user_id)

    # 첫 번째 입력 행 추가
    add_button = tk.Button(scrollable_frame , text="행 추가", command=lambda: add_entry_row(scrollable_frame , labels, session_data, len(session_data) + 1, add_button, save_button, delete_button, session_window_widgets=session_window_widgets))
    save_button = tk.Button(scrollable_frame , text="저장", command=lambda: save_session_data(window, session_window , {k: [e.get() for e in v] for k, v in session_data.items()}, user_id=user_id, user_data = user_data))
    delete_button = tk.Button(scrollable_frame , text="행 삭제", command=lambda: delete_entry_row(scrollable_frame , labels, session_data, add_button, save_button, delete_button, session_window_widgets=session_window_widgets))

    #기존 데이터를 사용해 입력 행 추가
    if user_data:
        for session in user_data:
            session_number = session[1]
            session_date = convert_date_format(session[2], slash_type)
            details = session[3]
            add_entry_row(scrollable_frame , labels, session_data, session_number, add_button, save_button, delete_button, preset_data=[session_date, details], session_window_widgets=session_window_widgets)
    else:
        add_entry_row(scrollable_frame , labels, session_data, 1, add_button, save_button, delete_button, session_window_widgets=session_window_widgets)
    
    # 메뉴바 생성
    update_font_size(settings["sessions_font_size"], session_window_widgets, sessions_bar)
    create_menu(session_window, session_window_widgets, sessions_bar)

def create_field_entries(window, user_id=None):
    """필드 및 라벨 설정을 함수화하여 중복 제거."""
    labels_and_fields = {}
    special_entries = {}
    window_widgets = []
    
    def calculate_age():
        try:
            """생년월일을 입력받아 나이를 계산합니다."""
            year_entry = labels_and_fields.get("년")
            month_entry = labels_and_fields.get("월")
            day_entry = labels_and_fields.get("일")

            birth_year = int(year_entry.get())
            birth_month = int(month_entry.get())
            birth_day = int(day_entry.get())
            birth_date = datetime(birth_year, birth_month, birth_day)

            today = datetime.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
            
            age_entry = labels_and_fields.get("나이")
            if age_entry:
                age_entry.config(state=tk.NORMAL)  # Enable editing
                age_entry.delete(0, tk.END)
                age_entry.insert(0, "-" if age == 0 else str(age))
                age_entry.config(state=tk.DISABLED)  # Disable editing
        except ValueError:
            pass  # 값이 입력되기 전에 발생하는 오류를 무시
    
    # Helper function to update date entry widget
    def on_key_release(event):
        # 모든 필드가 채워지면 나이 계산
        if all(labels_and_fields.get(field).get() for field in ["년", "월", "일"]):
            calculate_age()
    
    for header in headers:
        if header in ["ID"]:
            continue
        elif header in ["회기 수"]:
            entry = tk.Entry(window, width=10, validate="key", validatecommand=(window.register(validate_integer), "%P"))
            if user_id:
                session_button = tk.Button(window, text="상세", command=lambda: open_sessions(window, user_id))
                special_entries[header] = session_button
                
                window_widgets.append(session_button)
        elif header in ["상담시작일", "시작연도", "상담종료일", "종료연도"]:
            entry = tk.Entry(window, width=10, validate="key", validatecommand=(window.register(validate_integer), "%P"))
        elif header in ["시작월", "시작일", "종료월", "종료일"]:
            entry = tk.Entry(window, width=5, validate="key", validatecommand=(window.register(validate_integer), "%P"))
        elif header in ["전화번호"]:
            entry = tk.Entry(window, width=20, validate="key", validatecommand=(window.register(validate_phone), "%P"))
        elif header in ["생년월일", "년", "월", "일", "나이"]:
            # Create lists for year, month, and day dropdowns
            if header in ["년"]:
                entry = tk.Entry(window, width=10, validate="key", validatecommand=(window.register(validate_integer), "%P"))
                entry.bind('<KeyRelease>', on_key_release)
            elif header in ["월"]:
                entry = tk.Entry(window, width=5, validate="key", validatecommand=(window.register(validate_integer), "%P"))
                entry.bind('<KeyRelease>', on_key_release)
            elif header in ["일"]:
                entry = tk.Entry(window, width=5, validate="key", validatecommand=(window.register(validate_integer), "%P"))
                entry.bind('<KeyRelease>', on_key_release)
            elif header in ["생년월일"]:
                # Entry widget for date
                entry = tk.Entry(window, width=5)
            else:
                entry = tk.Entry(window, width=5, validate="key", validatecommand=(window.register(validate_integer), "%P"))
                entry.config(state=tk.DISABLED)  # Disable editing
        elif header in ["특이사항"]:
            entry = tk.Text(window, height=10, width=60)
        elif header in ["호소문제"]:
            entry = tk.Text(window, height=3, width=60)
        elif header in ["메일"]:
            entry = tk.Entry(window, width=30, validate="key", validatecommand=(window.register(validate_hyphen), "%P"))
        elif header in ["주소"]:
            entry = tk.Entry(window, width=60, validate="key", validatecommand=(window.register(validate_hyphen), "%P"))
        elif header in ["성별", "남자", "여자"]:
            if header in ["남자"]:
                entry = tk.Radiobutton(window, text="남", variable=gender_var, value="남")
            elif header in ["여자"]:
                entry = tk.Radiobutton(window, text="여", variable=gender_var, value="여")
            else:
                entry = tk.Entry(window, width=10)
            gender_var.set("-")
        else:
            entry = tk.Entry(window, width=10, validate="key", validatecommand=(window.register(validate_hyphen), "%P"))
        
        labels_and_fields[header] = entry
        window_widgets.append(entry)
    
    return labels_and_fields, special_entries, window_widgets

def grid_field_entries(labels_and_fields, special_entries, window, window_widgets, user_id=None):
    """필드 및 라벨의 그리드를 설정."""
    row = 0
    for label_text, entry in labels_and_fields.items():
        if label_text == "특이사항" or label_text == "호소문제":
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=0, padx=10, pady=5, sticky='ew')
            scrollbar = tk.Scrollbar(window, command=entry.yview)
            entry.config(yscrollcommand=scrollbar.set)

            # 특이사항 필드의 프레임과 텍스트 및 스크롤 바를 배치
            column_length = 20
            entry.grid(row=row, column=1, columnspan=column_length, padx=1, pady=5, sticky="w")
            scrollbar.grid(row=row, column=column_length+1, pady=5, sticky="ns")
            row += 1
        elif label_text == "생년월일" or label_text == "성별" or label_text == "상담시작일" or label_text == "상담종료일":
            # Handling the special case for "생년월일", "성별", "상담시작일", "상담종료일"
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=0, padx=10, pady=5, sticky='ew')
        elif label_text == "년":
            entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=2, padx=1, pady=5, sticky='w')
        elif label_text == "월":
            entry.grid(row=row, column=3, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=4, padx=1, pady=5, sticky='w')
        elif label_text == "일":
            entry.grid(row=row, column=5, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=6, padx=1, pady=5, sticky='w')
            row += 1
        elif label_text == "남자":
            entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')
        elif label_text == "여자":
            entry.grid(row=row, column=2, padx=1, pady=5, sticky='w')
            row += 1
        elif label_text == "시작연도":
            entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')
            label = tk.Label(window, text="년")
            label.grid(row=row, column=2, padx=1, pady=5, sticky='w')
        elif label_text == "시작월":
            entry.grid(row=row, column=3, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text="월")
            label.grid(row=row, column=4, padx=1, pady=5, sticky='w')
        elif label_text == "시작일":
            entry.grid(row=row, column=5, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text="일")
            label.grid(row=row, column=6, padx=1, pady=5, sticky='w')
            row += 1
        elif label_text == "종료연도":
            entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')
            label = tk.Label(window, text="년")
            label.grid(row=row, column=2, padx=1, pady=5, sticky='w')
        elif label_text == "종료월":
            entry.grid(row=row, column=3, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text="월")
            label.grid(row=row, column=4, padx=1, pady=5, sticky='w')
        elif label_text == "종료일":
            entry.grid(row=row, column=5, padx=1, pady=5, sticky='e')
            label = tk.Label(window, text="일")
            label.grid(row=row, column=6, padx=1, pady=5, sticky='w')
            row += 1
        else:
            label = tk.Label(window, text=label_text)
            label.grid(row=row, column=0, padx=10, pady=5, sticky='ew')
            if label_text == "이름" or label_text == "나이":
                entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')
                row += 1
            elif label_text == "회기 수":
                entry.grid(row=row, column=1, padx=1, pady=5, sticky='w')

                if user_id:
                    session_button = special_entries[label_text]
                    session_button.grid(row=row, column=2, padx=1, pady=5, sticky='w')
                row += 1
            else:
                entry.grid(row=row, column=1, columnspan=20, padx=1, pady=5, sticky='w')
                row += 1
        
        window_widgets.append(label)
    
    return window_widgets

def populate_fields(entries, user_data):
    """기존 데이터를 필드에 채워 넣음."""
    # 딕셔너리 항목을 리스트로 변환
    items = list(entries.items())
    
    for label, entry in items[0:]:
        index = headers.index(label) + 0  # ID가 첫 번째 필드이므로 +1
            # tk.Text, tk.Entry를, tk.Radiobutton 구분하여 처리
        if isinstance(entry, tk.Text):
            entry.delete("1.0", tk.END)  # 기존 내용을 삭제
            entry.insert("1.0", user_data[index])  # 텍스트 삽입
        elif isinstance(entry, tk.Radiobutton):
            if user_data[index] == "o":
                if label in ["남자"]:
                    gender_var.set("남")  # 남성 선택
                elif label in ["여자"]:
                    gender_var.set("여")  # 여성 선택
        elif isinstance(entry, tk.Button):
            pass
        else:
            if label in ["나이"]:
                entry.config(state=tk.NORMAL)  # Enable editing
                entry.delete(0, tk.END)  # 기존 내용을 삭제
                entry.insert(0, user_data[index])  # 텍스트 삽입
                entry.config(state=tk.DISABLED) # Disable editing
            else:
                entry.delete(0, tk.END)  # 기존 내용을 삭제
                entry.insert(0, user_data[index])  # 텍스트 삽입

def save_user_data(values, user_id=None):
    """사용자 데이터를 저장 또는 업데이트."""
    try:
        # 상담시작/종료일 병합
        start_year = values.get("시작연도", "").strip()
        start_month = values.get("시작월", "").strip()
        start_day = values.get("시작일", "").strip()
        end_year = values.get("종료연도", "").strip()
        end_month = values.get("종료월", "").strip()
        end_day = values.get("종료일", "").strip()
        
        # 상담시작일 기본값 설정
        if not start_year or not start_month or not start_day:
            consultation_start = "-"
        else:
            try:
                start_year = int(start_year)
                start_month = int(start_month)
                start_day = int(start_day)
                consultation_start = f"{start_year}-{start_month:02d}-{start_day:02d}"
            except ValueError:
                consultation_start = "-"  # 변환 실패 시 기본값 설정
        
        # 상담종료일 기본값 설정
        if not end_year or not end_month or not end_day:
            consultation_end = "-"
        else:
            try:
                end_year = int(end_year)
                end_month = int(end_month)
                end_day = int(end_day)
                consultation_end = f"{end_year}-{end_month:02d}-{end_day:02d}"
            except ValueError:
                consultation_end = "-"  # 변환 실패 시 기본값 설정
        
        # 생년월일 병합
        year = values.get("년", "").strip()
        month = values.get("월", "").strip()
        day = values.get("일", "").strip()
        
        # 기본값 설정
        if not year or not month or not day:
            birth = "-"
            age = "-"
        else:
            # 숫자로 변환하고 예외 처리
            age = values.get("나이", 0)
            try:
                year = int(year)
                month = int(month)
                day = int(day)
                birth = f"{year}-{month:02d}-{day:02d}"
            except ValueError:
                birth = "-"  # 변환 실패 시 기본값 설정
        
        #성별
        gender = gender_var.get()
        male = "o" if gender == "남" else ""
        female = "o" if gender == "여" else ""

        # 문자열이 숫자일 경우에만 정수로 변환
        value_sessions = values.get("회기 수", 0)
        
        # 나이, 회기 수 정수값여부 체크
        if value_sessions.isdigit():
            sessions = int(value_sessions)
        else:
            sessions = "-"  # 기본값
        
        if values["이름"]:
            if user_id:
                sql.update_user(
                    user_id,
                    name=values.get("이름", "") or "-",
                    birth=birth,
                    year=year,
                    month=month,
                    day=day,
                    age=age,
                    email=values.get("메일", "") or "-",
                    gender=gender_var.get() or "-",
                    male=male,
                    female=female,
                    phone=values.get("전화번호", "") or "-",
                    address=values.get("주소", "") or "-",
                    consultation_start=consultation_start,
                    start_year = start_year,
                    start_month = start_month,
                    start_day = start_day,
                    consultation_end=consultation_end,
                    end_year = end_year,
                    end_month = end_month,
                    end_day = end_day,
                    issue=values.get("호소문제", ""),
                    sessions=sessions,
                    notes=values.get("특이사항", "")
                )
                messagebox.showinfo("편집 완료", "리스트가 수정되었습니다.")
            else:
                sql.create_user(
                    name=values.get("이름", "") or "-",
                    birth=birth,
                    year=year,
                    month=month,
                    day=day,
                    age=age,
                    email=values.get("메일", "") or "-",
                    gender=gender_var.get() or "-",
                    male=male,
                    female=female,
                    phone=values.get("전화번호", "") or "-",
                    address=values.get("주소", "") or "-",
                    consultation_start=consultation_start,
                    start_year = start_year,
                    start_month = start_month,
                    start_day = start_day,
                    consultation_end=consultation_end,
                    end_year = end_year,
                    end_month = end_month,
                    end_day = end_day,
                    issue=values.get("호소문제", ""),
                    sessions=sessions,
                    notes=values.get("특이사항", "")
                )
                messagebox.showinfo("생성 완료", "생성되었습니다.")
            read_users_gui()
        else:
            messagebox.showwarning("Input Error", "올바른 항목을 입력하세요.")
    except Exception as e:
        messagebox.showerror("Error", f"오류가 발생했습니다: {e}")

def open_create_user_window():
    settings = load_settings()

    create_window = tk.Toplevel(root)
    create_window.title("신규 생성")
    create_window.geometry(settings["edit_window_size"])  # 새 창 크기 설정

    entries, special_entries, window_widgets = create_field_entries(create_window)
    window_widgets = grid_field_entries(entries, special_entries, create_window, window_widgets)
    
	# 새 창 열릴 때 Name 입력 필드에 포커스
    entry_name = entries["이름"]
    entry_name.focus_set()
    
    def save_new_user():
        values = {
            label: (entry.get() if isinstance(entry, tk.Entry) else 
            entry.get("1.0", tk.END).strip() if isinstance(entry, tk.Text) else 
            entry.get_date() if isinstance(entry, DateEntry) else None)
            for label, entry in entries.items()
        }
        save_user_data(values)
        create_window.destroy()
    
    button_save = tk.Button(create_window, text="저장", command=save_new_user)
    button_save.grid(row=len(headers), column=20, padx=5, pady=5)
    
    # 메뉴바 생성
    update_font_size(settings["edit_font_size"], window_widgets, edit_bar)
    create_menu(create_window, window_widgets, edit_bar)

def open_update_window(user_id):
    settings = load_settings()

    user_data = [row for row in sql.read_users() if str(row[0]) == user_id][0]
    update_window = tk.Toplevel(root)
    update_window.title("편집")
    update_window.geometry(settings["edit_window_size"])  # 새 창 크기 설정

    entries, special_entries, window_widgets = create_field_entries(update_window, user_id)
    window_widgets = grid_field_entries(entries, special_entries, update_window, window_widgets, user_id)

    # 기존 데이터로 필드 채우기
    populate_fields(entries, user_data)
    
	# 새 창 열릴 때 Name 입력 필드에 포커스
    entry_name = entries["이름"]
    entry_name.focus_set()
    
    def save_updates():
        values = {
            label: (entry.get() if isinstance(entry, tk.Entry) else 
            entry.get("1.0", tk.END).strip() if isinstance(entry, tk.Text) else 
            entry.get_date() if isinstance(entry, DateEntry) else None)
            for label, entry in entries.items()
        }
        save_user_data(values, user_id=user_id)
        update_window.destroy()
    
    button_export = tk.Button(update_window, text="상세 출력", command=lambda: export_to_excel(user_id=user_id))
    button_export.grid(row=0, column=20, padx=5, pady=5)

    button_save = tk.Button(update_window, text="저장", command=save_updates)
    button_save.grid(row=len(headers), column=20, padx=5, pady=5)
    
    window_widgets.append(button_export)
    window_widgets.append(button_save)

    # 메뉴바 생성
    update_font_size(settings["edit_font_size"], window_widgets, edit_bar)
    create_menu(update_window, window_widgets, edit_bar)

def read_users_gui(search_query=None):
    treeview_users.delete(*treeview_users.get_children())
    rows = sql.read_users(search_query)  # 검색 쿼리를 전달
    
    for row in rows:
        formatted_row = list(row)
        if not formatted_row[3] or not formatted_row[4] or not formatted_row[5]:
            age = "-"
        else:
            age = update_age(formatted_row[3], formatted_row[4], formatted_row[5])
        sql.update_user(
                            formatted_row[0],
                            name=formatted_row[1],
                            birth=formatted_row[2],
                            year=formatted_row[3],
                            month=formatted_row[4],
                            day=formatted_row[5],
                            age=age,
                            email=formatted_row[7],
                            gender=formatted_row[8],
                            male=formatted_row[9],
                            female=formatted_row[10],
                            phone=formatted_row[11],
                            address=formatted_row[12],
                            consultation_start=formatted_row[13],
                            start_year=formatted_row[14],
                            start_month=formatted_row[15],
                            start_day=formatted_row[16],
                            consultation_end=formatted_row[17],
                            end_year=formatted_row[18],
                            end_month=formatted_row[19],
                            end_day=formatted_row[20],
                            issue=formatted_row[21],
                            sessions=formatted_row[22],
                            notes=formatted_row[23]
                        )
    
    rows = sql.read_users(search_query)  # 검색 쿼리를 전달
    
    for row in rows:
        formatted_row = list(row)
        for i, header in enumerate(headers):
            value = row[i]
            if isinstance(value, str):
                formatted_row[i] = value
            elif isinstance(value, int):
                formatted_row[i] = str(value)
        
        # headers 리스트에서 viewonly_headers에 해당하는 항목의 인덱스하여 selected_column에 저장
        indices = [headers.index(header) for header in viewonly_headers]
        selected_column = [formatted_row[i] for i in indices]

        treeview_users.insert("", tk.END, iid=str(row[0]), values=selected_column)
        
    # 메뉴바 생성
    main_widgets = [label_search, option_menu, entry_search, button_search, button_show_all, button_create, button_update, button_delete, export_button]
    
    # 메뉴바 생성
    settings = load_settings()
    update_font_size(settings["main_font_size"], main_widgets, main_bar)
    create_menu(root, main_widgets, main_bar)

def on_user_select(event=None):
    selected_item = treeview_users.selection()
    if selected_item:
        user_id = selected_item[0]
        open_update_window(user_id)

def delete_user_gui():
    selected_item = treeview_users.selection()
    
    if selected_item:
        user_names = []
        user_ids = []
        for item in selected_item:
            user_data = treeview_users.item(item, 'values')
            user_name = user_data[1]
            user_id = item  # user_id를 item에서 추출

            user_names.append(user_name)
            user_ids.append(user_id)

        # 선택된 모든 유저를 한 번에 삭제할지 확인
        confirm = messagebox.askyesno("리스트 삭제", f"{', '.join(user_names)}을 삭제하시겠습니까?")
        
        if confirm:
            for user_id in user_ids:
                sql.delete_user(user_id)
            messagebox.showinfo("삭제 완료", "리스트가 삭제되었습니다.")
            read_users_gui()
    else:
        messagebox.showwarning("Selection Error", "Please select a user from the list.")

def search_users(event=None):
    search_field = search_field_var.get()
    search_value = entry_search.get()
    search_query = {search_field: search_value}
    read_users_gui(search_query)

def show_all_users():
    entry_search.delete(0, tk.END)  # 검색 입력 필드 비우기
    read_users_gui()  # 전체 목록 불러오기

def export_to_excel(user_id=None):
    """현재 데이터를 엑셀 파일로 내보내는 함수"""
    if user_id is None:
        data = sql.fetch_users()  # 데이터베이스에서 고객 데이터 가져오기
        file_name = "고객_목록.xlsx"
        columns_list = viewonly_headers
        
        df = pd.DataFrame(data, columns=columns_list)
        df.to_excel(file_name, index=False, engine='openpyxl')
        messagebox.showinfo("정보", "고객 목록 엑셀 파일로 내보내졌습니다.")
    else:
        user_data = sql.fetch_users_by_id(user_id)  # 데이터베이스에서 고객 데이터 가져오기
        user_name = sql.fetch_user_name_by_id(user_id)
        
        # 세션 세부 정보 가져오기
        session_details = sql.fetch_session_details_by_user(user_id, export_detail_headers)
        
        data = list(user_data)
        for session in session_details:
            combined_data = data + list(session)
            # 중첩된 튜플을 리스트로 풀어줌
            flattened_data = [item for sublist in combined_data for item in (sublist if isinstance(sublist, tuple) else [sublist])]
            data.append(flattened_data)
        
        file_name = user_name + "_정보.xlsx"

        columns_list = export_headers + export_detail_headers
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f"{user_name}님 정보"
        
        # 황갈색 색상 채우기
        light_gray_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

        # 첫 번째 행: 고객정보
        ws.merge_cells('A1:J1')
        ws['A1'] = "고 객 정 보"
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].fill = light_gray_fill
        
        # 두 번째 행: 이름, 상담 시작일, 상담 종료일
        ws.merge_cells('B2:D2')
        ws.merge_cells('F2:G2')
        ws.merge_cells('I2:J2')
        ws['A2'] = "이 름"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(bold=True)
        ws['A2'].fill = light_gray_fill
        ws['B2'] = data[0][1]
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['E2'] = "상담 시작일"
        ws['E2'].alignment = Alignment(horizontal='center')
        ws['E2'].font = Font(bold=True)
        ws['E2'].fill = light_gray_fill
        ws['F2'] = data[0][8]
        ws['F2'].alignment = Alignment(horizontal='center')
        ws['H2'] = "상담 종료일"
        ws['H2'].alignment = Alignment(horizontal='center')
        ws['H2'].font = Font(bold=True)
        ws['H2'].fill = light_gray_fill
        ws['I2'] = data[0][9]
        ws['I2'].alignment = Alignment(horizontal='center')

        # 세 번째 행: 성별, 나이, 생년월일
        ws.merge_cells('F3:G3')
        ws.merge_cells('H3:J3')
        ws['A3'] = "성 별"
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = light_gray_fill
        ws['B3'] = data[0][5]
        ws['B3'].alignment = Alignment(horizontal='center')
        ws['C3'] = "나 이"
        ws['C3'].alignment = Alignment(horizontal='center')
        ws['C3'].font = Font(bold=True)
        ws['C3'].fill = light_gray_fill
        ws['D3'] = data[0][3]
        ws['D3'].alignment = Alignment(horizontal='center')
        ws['E3'] = "생년월일"
        ws['E3'].alignment = Alignment(horizontal='center')
        ws['E3'].font = Font(bold=True)
        ws['E3'].fill = light_gray_fill
        ws['F3'] = data[0][2]
        ws['F3'].alignment = Alignment(horizontal='center')

        # 네 번째 행: 연락처, 메일
        ws.merge_cells('B4:D4')
        ws.merge_cells('F4:J4')
        ws['A4'] = "연 락 처"
        ws['A4'].alignment = Alignment(horizontal='center')
        ws['A4'].font = Font(bold=True)
        ws['A4'].fill = light_gray_fill
        ws['B4'] = data[0][6]
        ws['B4'].alignment = Alignment(horizontal='center')
        ws['E4'] = "메 일"
        ws['E4'].alignment = Alignment(horizontal='center')
        ws['E4'].font = Font(bold=True)
        ws['E4'].fill = light_gray_fill
        ws['F4'] = data[0][4]
        ws['F4'].alignment = Alignment(horizontal='left')
        
        # 다섯 번째 행: 주소
        ws.merge_cells('B5:J5')
        ws['A5'] = "주 소"
        ws['A5'].alignment = Alignment(horizontal='center')
        ws['A5'].font = Font(bold=True)
        ws['A5'].fill = light_gray_fill
        ws['B5'] = data[0][7]
        ws['B5'].alignment = Alignment(horizontal='left')

        # 여섯 번째 행: 회기수, 호소문제
        ws.merge_cells('D6:J6')
        ws['A6'] = "회기 수"
        ws['A6'].alignment = Alignment(horizontal='center')
        ws['A6'].font = Font(bold=True)
        ws['A6'].fill = light_gray_fill
        ws['B6'] = data[0][11]
        ws['B6'].alignment = Alignment(horizontal='center')
        ws['C6'] = "호소 문제"
        ws['C6'].alignment = Alignment(horizontal='center')
        ws['C6'].font = Font(bold=True)
        ws['C6'].fill = light_gray_fill
        ws['D6'] = data[0][10]
        ws['D6'].alignment = Alignment(horizontal='left')
        
        # 일곱, 여덟 번째 행: 특이사항
        ws.merge_cells('A7:J7')
        ws.merge_cells('A8:J18')
        ws['A7'] = "특 이 사 항"
        ws['A7'].alignment = Alignment(horizontal='center')
        ws['A7'].font = Font(bold=True, size=13)
        ws['A7'].fill = light_gray_fill
        ws['A8'] = data[0][12]
        ws['A8'].alignment = Alignment(horizontal='left', vertical='top')

        # 열아홉, 스물 번째 행: 세션 정보 헤더
        ws.merge_cells('A19:J19')
        ws.merge_cells('B20:C20')
        ws.merge_cells('D20:J20')
        ws['A19'] = "상 세"
        ws['A19'].alignment = Alignment(horizontal='center')
        ws['A19'].font = Font(bold=True, size=13)
        ws['A19'].fill = light_gray_fill
        ws['A20'] = "회 차"
        ws['A20'].alignment = Alignment(horizontal='center')
        ws['A20'].font = Font(bold=True)
        ws['A20'].fill = light_gray_fill
        ws['B20'] = "날 짜"
        ws['B20'].alignment = Alignment(horizontal='center')
        ws['B20'].font = Font(bold=True)
        ws['B20'].fill = light_gray_fill
        ws['D20'] = "상담 내용"
        ws['D20'].alignment = Alignment(horizontal='center')
        ws['D20'].font = Font(bold=True)
        ws['D20'].fill = light_gray_fill
        
        max_row = 20
        # 열 번째 행부터: 세션 데이터 입력
        for idx, session in enumerate(session_details, start=21):
            ws.merge_cells(f'B{idx}:C{idx}')
            ws.merge_cells(f'D{idx}:J{idx}')
            ws[f'A{idx}'] = session[0]  # 회차 번호
            ws[f'A{idx}'].alignment = Alignment(horizontal='center')
            ws[f'B{idx}'] = session[1]  # 날짜
            ws[f'B{idx}'].alignment = Alignment(horizontal='center')
            ws[f'D{idx}'] = session[2]  # 내용
            ws[f'D{idx}'].alignment = Alignment(horizontal='left')
            max_row = idx
        
        # 테두리 스타일 설정
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 고객 정보 테이블 셀에 테두리 적용
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=10):
            for cell in row:
                if cell.column == 1:
                    if cell.row == 1:
                        cell.border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                    elif cell.row == max_row:
                        cell.border = Border(left=Side(style='medium'), bottom=Side(style='medium'))
                    elif cell.row == 7 or cell.row == 19:
                        cell.border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                    else:
                        cell.border = Border(left=Side(style='medium'), bottom=Side(style='thin'))
                elif cell.column == 10:
                    if cell.row == 1:
                        cell.border = Border(right=Side(style='medium'),top=Side(style='medium'), bottom=Side(style='medium'))
                    elif cell.row == max_row:
                        cell.border = Border(right=Side(style='medium'), bottom=Side(style='medium'))
                    elif cell.row == 7 or cell.row == 19:
                        cell.border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                    else:
                        cell.border = Border(right=Side(style='medium'), bottom=Side(style='thin'))
                elif cell.row == 1:
                    cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
                elif cell.row == 7 or cell.row == 19:
                    cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
                elif cell.row == max_row:
                    cell.border = Border(left=Side(style='thin'),bottom=Side(style='medium'))
                else:
                    cell.border = thin_border
        
        # 셀 너비 조정 함수
        def auto_adjust_column_width(ws, min_width=12):
            for col_index, col in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_index)  # 열 인덱스에서 열 레터 얻기
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 4)  # Add a little extra space
                ws.column_dimensions[column_letter].width = max(adjusted_width, min_width)

        # 자동 셀 너비 조정
        auto_adjust_column_width(ws)

        # 파일 이름 설정
        file_name = f"{user_name}_정보.xlsx"
        wb.save(file_name)
        messagebox.showinfo("정보", user_name + "님 정보가 엑셀 파일로 내보내졌습니다.")
    
def on_closing():
    save_settings(settings)
    root.destroy()

search_frame = tk.Frame(root)
search_frame.grid(row=0, column=0, columnspan=5, pady=1, sticky='w')  # 상단 좌측에 배치

# 검색 필드 및 드롭다운 메뉴 설정
label_search = tk.Label(search_frame, text="분류")
label_search.grid(row=0, column=0, padx=10, pady=10, sticky='e')

search_field_var = tk.StringVar(value="이름")  # 기본 검색 필드를 이름으로 설정

# 드롭다운 메뉴 (Combobox)로 필드 선택
option_menu = ttk.Combobox(search_frame, textvariable=search_field_var, values=viewonly_headers, width=20)
option_menu.grid(row=0, column=1, padx=10, pady=10, sticky='w')

# 검색 입력 필드
entry_search = tk.Entry(search_frame, width=30)
entry_search.grid(row=0, column=2, padx=10, pady=10, sticky='w')

# 검색 버튼
button_search = tk.Button(search_frame, text="검색", command=search_users)
button_search.grid(row=0, column=3, padx=10, pady=10, sticky='w')

# 전체 목록 보기 버튼
button_show_all = tk.Button(search_frame, text="전체 보기", command=show_all_users)
button_show_all.grid(row=0, column=4, padx=10, pady=10, sticky='w')

# 버튼들을 1행으로 배치하기 위한 Frame 사용
button_frame = tk.Frame(root)
button_frame.grid(row=1, column=0, columnspan=5, pady=1, sticky='e')  # 상단 우측에 배치

button_create = tk.Button(button_frame, text="신규 생성", command=open_create_user_window)
button_create.grid(row=0, column=0, padx=10)

button_update = tk.Button(button_frame, text="편집", command=on_user_select)
button_update.grid(row=0, column=1, padx=10)

button_delete = tk.Button(button_frame, text="삭제", command=delete_user_gui)
button_delete.grid(row=0, column=2, padx=10)

export_button = tk.Button(button_frame, text="목록 출력", command=export_to_excel)
export_button.grid(row=0, column=3, padx=5, pady=5)

# Bind the Enter key to the search_users function
entry_search.bind("<Return>", search_users)

# Grid 레이아웃의 행과 열 비율 조정
root.grid_rowconfigure(2, weight=1)
root.grid_columnconfigure(2, weight=1)

# Treeview의 더블 클릭 이벤트 바인딩
treeview_users.bind("<Double-1>", on_user_select)

# 시작 시 사용자 목록 읽기
read_users_gui()

# 스크롤바 추가
scrollbar = tk.Scrollbar(root, orient="vertical", command=treeview_users.yview)
scrollbar.grid(row=2, column=5, sticky="ns")
treeview_users.configure(yscrollcommand=scrollbar.set)

# config 저장
root.protocol("WM_DELETE_WINDOW", on_closing)

# GUI 루프 시작
root.mainloop()

# 프로그램 종료 시 데이터베이스 연결 종료
sql.close_connection()